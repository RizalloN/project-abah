import argparse
import csv
import json
import os
import sys
from datetime import date, datetime


HEADERS = [
    "periode",
    "billing",
    "kanca",
    "bc",
    "mbm",
    "uker",
    "pn",
    "mantri",
    "no_rekening",
    "nama_debitur",
    "plafon",
    "next_pmt_date",
    "update_npd",
    "tgl_realisasi",
    "tgl_jatuh_tempo",
    "jangka_waktu",
    "flag_restruk",
    "m_min_1_kol",
    "m_min_1_detail",
    "m_min_1_os",
    "wba",
    "now_kol",
    "now_detail",
    "now_os",
    "now_t_pokok",
    "now_t_bunga",
    "now_t_total",
    "ptp",
]

DATE_INDEXES = {0, 11, 12, 13, 14}
UPDATE_NPD_INDEX = 12
NOW_KOL_INDEX = 21
NOW_DETAIL_INDEX = 22
NOW_OS_INDEX = 23
NOW_AMOUNT_INDEXES = {23, 24, 25, 26}


def emit(payload):
    print(json.dumps(payload, ensure_ascii=False), flush=True)


def cell_to_csv_value(value, index):
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y") if index in DATE_INDEXES else value.isoformat(sep=" ")

    if isinstance(value, date):
        return value.strftime("%d/%m/%Y") if index in DATE_INDEXES else value.isoformat()

    return str(value).strip()


def looks_like_npd_header(row):
    values = [str(value or "").strip().upper() for value in row]

    return (
        "BILLING" in values
        and "PERIODE" in values
        and "KANCA" in values
        and "NOMOR REKENING" in values
        and "UPDATE NPD" in values
    )


def normalize_lunas_row(values):
    normalized = list(values)
    is_lunas = any(
        str(normalized[index]).strip().upper() == "LUNAS"
        for index in (UPDATE_NPD_INDEX, NOW_KOL_INDEX, NOW_DETAIL_INDEX)
        if index < len(normalized)
    )

    if not is_lunas:
        return normalized

    if UPDATE_NPD_INDEX < len(normalized):
        normalized[UPDATE_NPD_INDEX] = "LUNAS"
    if NOW_KOL_INDEX < len(normalized):
        normalized[NOW_KOL_INDEX] = "Lunas"
    if NOW_DETAIL_INDEX < len(normalized):
        normalized[NOW_DETAIL_INDEX] = "Lunas"

    for index in NOW_AMOUNT_INDEXES:
        if index < len(normalized):
            normalized[index] = "0"

    return normalized


def main():
    parser = argparse.ArgumentParser(description="Stream LW321 NPD MICRO XLSX to normalized CSV staging.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output")
    parser.add_argument("--preview-only", action="store_true")
    parser.add_argument("--preview-limit", type=int, default=75)
    parser.add_argument("--progress-every", type=int, default=25000)
    args = parser.parse_args()

    if not args.preview_only and not args.output:
        raise RuntimeError("--output wajib diisi untuk mode staging penuh.")

    if args.output:
        os.makedirs(os.path.dirname(args.output), exist_ok=True)

    from openpyxl import load_workbook

    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    worksheet_max_row = ws.max_row or 0

    header_row = None
    rows_written = 0
    preview_rows = []
    unique_values = {}

    try:
        handle = None
        writer = None
        if not args.preview_only:
            handle = open(args.output, "w", newline="", encoding="utf-8")
            writer = csv.writer(handle)
            writer.writerow(HEADERS)

        try:
            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = list(row)

                if header_row is None:
                    if looks_like_npd_header(values):
                        header_row = row_number
                        emit({
                            "type": "progress",
                            "percent": 45 if not args.preview_only else 65,
                            "message": "Header LW321 NPD ditemukan. Menulis CSV staging..." if not args.preview_only else "Header LW321 NPD ditemukan. Menyiapkan sampel preview...",
                            "header_row": header_row,
                            "headers": len(HEADERS),
                        })
                    continue

                if row_number == header_row + 1:
                    continue

                normalized = [cell_to_csv_value(value, index) for index, value in enumerate((values + [""] * len(HEADERS))[:len(HEADERS)])]
                normalized = normalize_lunas_row(normalized)
                if not any(value for value in normalized):
                    continue

                if writer is not None:
                    writer.writerow(normalized)

                rows_written += 1

                if len(preview_rows) < args.preview_limit:
                    preview_rows.append(normalized)

                    for index, value in enumerate(normalized):
                        value = str(value).strip() or "(Blank)"
                        bucket = unique_values.setdefault(index, {})
                        if len(bucket) < 75:
                            bucket[value] = True

                if args.preview_only and len(preview_rows) >= args.preview_limit:
                    break

                if not args.preview_only and rows_written % max(1, args.progress_every) == 0:
                    emit({
                        "type": "progress",
                        "percent": min(78, 45 + int(rows_written / max(worksheet_max_row, 1) * 35)),
                        "message": f"Menulis CSV staging LW321 NPD... {rows_written} baris",
                        "rows": rows_written,
                    })
        finally:
            if handle is not None:
                handle.close()
    finally:
        wb.close()

    if header_row is None:
        raise RuntimeError("Header LW321 NPD tidak ditemukan. Pastikan file memuat Billing, KANCA, NOMOR REKENING, dan UPDATE NPD.")

    emit({
        "type": "done",
        "output": args.output if not args.preview_only else None,
        "header_index": 0,
        "source_header_row": header_row,
        "headers": HEADERS,
        "preview_rows": preview_rows,
        "unique_values": {str(index): list(values.keys()) for index, values in unique_values.items()},
        "total_rows": rows_written if not args.preview_only else max(0, worksheet_max_row - header_row - 1),
    })


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        emit({"type": "error", "message": str(exc)})
        sys.exit(1)
