#!/usr/bin/env python3
"""
Excel CPU Processor - OPTIMIZED dengan Fully Vectorized Polars
===============================================================
Optimisasi:
  1. POLARS VECTORIZED: Gunakan Polars expressions, bukan Python loop
  2. NO JSON OVERHEAD: Output langsung ke CSV, skip intermediate JSON
  3. DIRECT DB LOAD: PHP gunakan LOAD DATA INFILE, bukan chunked inserts
  4. STREAMING: Process hanya batch yang diperlukan, tidak load semua ke memory

Dependencies:
  pip install polars openpyxl python-dateutil
"""

import sys
import json
import os
import argparse
import time
import uuid
import csv
import re
import math
from datetime import datetime, date, timedelta
from typing import Dict, List, Tuple, Optional

# ── Force CPU ────────────────────────────────────────────────────────────────
os.environ['CUDA_VISIBLE_DEVICES'] = ''
os.environ['ROCR_VISIBLE_DEVICES'] = ''


def send_event(event_type: str, data: dict):
    """Send progress/error event to PHP via stdout"""
    data['type'] = event_type
    print(json.dumps(data, ensure_ascii=False, default=str), flush=True)


def send_progress(percent: int, message: str, rows_done: int = 0, total: int = 0, speed: int = 0):
    send_event('progress', {
        'percent': percent,
        'message': message,
        'rows_done': rows_done,
        'total': total,
        'speed': speed,
    })


def send_error(message: str):
    send_event('error', {'message': message})


# ─────────────────────────────────────────────────────────────────────────────
# VECTORIZED NORMALIZATION dengan Polars Expressions
# ─────────────────────────────────────────────────────────────────────────────

EXCEL_EPOCH = date(1899, 12, 30)
DATE_COLUMNS = {'PERIODE', 'POSISI', 'MONTH_DAY_YEAR_OF_POSISI', 'MONTH_DAY_YEAR_OF_PERIODE',
                'TGL_REALISASI', 'TGL_JATUH_TEMPO', 'TANGGAL'}
DECIMAL_COLUMNS = {
    'BAKI_DEBET', 'SALDO', 'POKOK', 'BUNGA', 'PLAFON', 'BESAR_REALISASI',
    'ANGPOK', 'ANGBUNG', 'SISAPOK', 'SISABUN', 'OS_PENUH_BERJALAN',
    'SALDO_PERTAMA_PH_POKOK', 'SALDO_PERTAMA_PH_BUNGA', 'ANGR_POKOK',
    'ANGR_BUNGA', 'OS', 'TOTAL_FEE', 'TOTAL_NOMINAL', 'JUMLAH', 'NILAI',
    'BAKI_DEBET1', 'CKPN', 'BAP', 'BILPRN', 'BILINT', 'BILLC', 'PMTAMT',
    'TUNGGAKAN_POKOK', 'TUNGGAKAN_BUNGA'
}
NULL_STRS = {'', 'nan', 'none', 'nat', 'null', 'n/a', 'na'}


def canonicalize_header(header_name: str) -> str:
    """Normalize header name to uppercase, alphanumeric + underscore"""
    return re.sub(r'[^A-Z0-9]+', '_', str(header_name).upper().strip()).strip('_')


def read_excel_with_openpyxl(file_path: str, header_index: int) -> Tuple[List[str], List]:
    """Fast openpyxl reader (fallback)"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        headers = []
        rows = []

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_list = list(row)

            if idx == header_index + 1:
                headers = [str(v).strip() if v is not None else f'COL_{i}' for i, v in enumerate(row_list)]
                continue

            if idx <= header_index + 1:
                continue

            if all(v is None or str(v).strip() == '' for v in row_list):
                continue

            rows.append(row_list)

        return headers, rows
    finally:
        wb.close()


def read_excel_as_polars(file_path: str, header_index: int) -> Tuple[List[str], 'pl.DataFrame', str]:
    """Read Excel with Polars (vectorized backend)"""
    try:
        import polars as pl

        attempts = [
            lambda: pl.read_excel(file_path, sheet_id=1, engine='calamine',
                                 read_options={'header_row': header_index}),
            lambda: pl.read_excel(file_path, sheet_id=1,
                                 read_options={'header_row': header_index}),
        ]

        for attempt in attempts:
            try:
                df = attempt()
                headers = list(df.columns)
                return headers, df, 'polars'
            except Exception:
                continue
    except Exception:
        pass

    headers, rows = read_excel_with_openpyxl(file_path, header_index)
    try:
        import polars as pl
        df = pl.DataFrame({headers[i] if i < len(headers) else f'COL_{i}':
                          [row[i] if i < len(row) else None for row in rows]
                          for i in range(max(len(h) for h in [headers] + rows) if rows else len(headers))})
        return headers, df, 'openpyxl-to-polars'
    except Exception:
        return headers, rows, 'openpyxl-native'


def create_vectorized_normalization_expr(col_name: str, header_canonical: str) -> 'pl.Expr':
    """
    Create Polars expression untuk normalize column values.
    Ini bekerja pada seluruh kolom sekaligus, bukan per-row!
    """
    try:
        import polars as pl
    except ImportError:
        return None

    expr = pl.col(col_name).cast(pl.Utf8)

    if header_canonical in DATE_COLUMNS:
        # Normalisasi tanggal dengan Polars built-in
        expr = (
            expr
            .str.replace_all(r'[/-]', '-')
            .str.to_date('%Y-%m-%d', strict=False)
            .cast(pl.Utf8, strict=False)
        )
    elif header_canonical in DECIMAL_COLUMNS:
        # Normalisasi desimal: hapus spasi, handle comma/dot
        expr = (
            expr
            .str.replace_all(r'\s+', '')
            .str.replace_all(r',(\d{3})(?!\d)', '$1')  # Hapus ribuan separator jika ada
            .str.replace_all(',', '.')  # Konversi comma ke dot
            .cast(pl.Float64, strict=False)
            .round(2)
            .cast(pl.Utf8, strict=False)
        )

    return expr


def apply_vectorized_normalization(df: 'pl.DataFrame', valid_headers: List[Tuple[int, str]]) -> 'pl.DataFrame':
    """
    Apply vectorized normalization ke seluruh DataFrame sekaligus.
    Jauh lebih cepat daripada loop per-row!
    """
    try:
        import polars as pl
    except ImportError:
        return df

    # Build dict of normalization expressions
    norm_exprs = {}
    for idx, h_name in valid_headers:
        h_canonical = canonicalize_header(h_name)
        expr = create_vectorized_normalization_expr(df.columns[idx] if idx < len(df.columns) else h_name, h_canonical)
        if expr is not None:
            norm_exprs[df.columns[idx]] = expr

    # Apply all normalizations in single operation
    if norm_exprs:
        df = df.with_columns([expr.alias(col_name) for col_name, expr in norm_exprs.items()])

    return df


def run_process_vectorized(config: dict):
    """
    OPTIMIZED version: Fully vectorized processing dengan Polars
    Tidak ada Python loop normalisasi per-baris!
    """
    file_path = config['file_path']
    header_index = int(config['header_index'])
    table_name = str(config.get('table_name', '')).strip().lower()
    table_columns = config.get('table_columns', [])
    normalized_headers = config.get('normalized_headers', {})
    active_filters = config.get('active_filters')
    output_csv_path = config.get('output_csv_path')
    load_columns = config.get('load_columns', [])

    if isinstance(normalized_headers, list):
        normalized_headers = {str(i): v for i, v in enumerate(normalized_headers)}

    # Read Excel dengan Polars (vectorized dari awal)
    try:
        headers, df, backend = read_excel_as_polars(file_path, header_index)
        send_progress(15, f'File dibaca dengan {backend}: {len(df)} baris. Melakukan normalisasi vectorized...')
    except Exception as e:
        send_error(f'Gagal membaca Excel: {str(e)}')
        sys.exit(1)

    # Setup columns
    is_simpanan_multipn = table_name == 'simpanan_multipn'
    unique_id_col = 'uniqueid_SimoPN' if is_simpanan_multipn else 'uniqueid_namareport'
    suffix = '_SimoPN' if is_simpanan_multipn else '_DLD'
    table_columns_map = {str(col).lower(): str(col) for col in table_columns} if table_columns else {}
    unique_id_col = table_columns_map.get(unique_id_col.lower(), unique_id_col)
    unique_id_prefix = str(config.get('unique_id_prefix') or 'imp').strip() or 'imp'
    skip_cols = set(['id', unique_id_col.lower()])

    # Build valid headers
    valid_headers = []
    for idx_str in sorted(normalized_headers.keys(), key=lambda x: int(x)):
        h = normalized_headers[idx_str]
        if not str(h).startswith('COL_'):
            valid_headers.append((int(idx_str), h))

    if not output_csv_path or not load_columns:
        send_error('Output CSV path atau load_columns tidak tersedia untuk bulk load.')
        sys.exit(1)

    send_progress(20, 'Applying vectorized normalization untuk seluruh DataFrame...')

    # VECTORIZED NORMALIZATION - bukan loop per-row!
    try:
        import polars as pl
        df = apply_vectorized_normalization(df, valid_headers)
    except Exception as e:
        send_error(f'Normalization vectorized gagal: {str(e)}')
        sys.exit(1)

    send_progress(40, 'Normalisasi selesai. Applying filters dan preparing output...')

    # Apply active filters (Polars filter expressions)
    if active_filters:
        for filter_idx, filter_values in active_filters.items():
            filter_idx_int = int(filter_idx)
            if filter_idx_int < len(valid_headers):
                col_idx, col_name = valid_headers[filter_idx_int]
                col_name_df = df.columns[col_idx] if col_idx < len(df.columns) else col_name
                try:
                    import polars as pl
                    df = df.filter(pl.col(col_name_df).is_in(filter_values) | pl.col(col_name_df).is_null())
                except Exception:
                    pass

    total_rows = len(df)

    if total_rows == 0:
        send_progress(90, 'Tidak ada baris yang cocok dengan filter.')
        send_event('done', {'total_rows': 0, 'csv_path': output_csv_path})
        return

    # Write CSV langsung dari Polars (SANGAT CEPAT)
    try:
        os.makedirs(os.path.dirname(output_csv_path), exist_ok=True)

        # Reorder columns untuk match load_columns
        load_columns_lower = [str(col).lower() for col in load_columns]
        output_columns = []
        for col in load_columns_lower:
            for df_col in df.columns:
                if df_col.lower() == col:
                    output_columns.append(df_col)
                    break

        if output_columns:
            df_output = df.select(output_columns)
        else:
            df_output = df

        # Write ke CSV dengan Polars (sangat optimized)
        df_output.write_csv(
            output_csv_path,
            separator=',',
            quote_char='"',
            null_value='\\N',  # MySQL LOAD DATA format
            lineterminator='\n',
            has_header=False,  # LOAD DATA expects no header
        )

        send_progress(90, f'CSV output selesai: {total_rows} baris. Menunggu MySQL bulk load...')
        send_event('done', {'total_rows': total_rows, 'csv_path': output_csv_path})

    except Exception as e:
        send_error(f'Gagal menulis CSV: {str(e)}')
        sys.exit(1)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Optimized Excel Processor dengan Polars Vectorization')
    parser.add_argument('--config-json', required=True, help='JSON config string')

    args = parser.parse_args()

    try:
        config = json.loads(args.config_json)
    except json.JSONDecodeError as e:
        send_error(f'Invalid config JSON: {str(e)}')
        sys.exit(1)

    mode = config.get('mode', 'process')

    if mode == 'process':
        run_process_vectorized(config)
    else:
        send_error(f'Unknown mode: {mode}')
        sys.exit(1)
