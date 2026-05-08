import argparse
import csv
import json
import os
import sys
from datetime import date, datetime


HEADERS = [
    "periode",
    "billing",
    "kanca",
    "bc",
    "mbm",
    "uker",
    "pn",
    "mantri",
    "no_rekening",
    "nama_debitur",
    "plafon",
    "npdd",
    "npdd_update",
    "tgl_realisasi",
    "tgl_jatuh_tempo",
    "jangka_waktu",
    "flag_restruk",
    "kol",
    "detail",
    "os",
    "wba",
    "now_kol",
    "now_detail",
    "now_os",
    "now_t_pokok",
    "now_t_bunga",
    "now_t_total",
    "ptp",
]

DATE_INDEXES = {0, 11, 12, 13, 14}
NPDD_UPDATE_INDEX = 12
NOW_KOL_INDEX = 21
NOW_DETAIL_INDEX = 22
NOW_NUMERIC_INDEXES = {23, 24, 25, 26}
PTP_INDEX = 27


def emit(payload):
    print(json.dumps(payload, ensure_ascii=False), flush=True)


def cell_to_csv_value(value, index):
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y") if index in DATE_INDEXES else value.isoformat(sep=" ")

    if isinstance(value, date):
        return value.strftime("%d/%m/%Y") if index in DATE_INDEXES else value.isoformat()

    return str(value).strip()


def normalize_lunas_row(values):
    normalized = list(values)
    now_kol = str(normalized[NOW_KOL_INDEX] if len(normalized) > NOW_KOL_INDEX else "").strip().upper()
    now_detail = str(normalized[NOW_DETAIL_INDEX] if len(normalized) > NOW_DETAIL_INDEX else "").strip().upper()

    if now_kol != "LUNAS" and now_detail != "LUNAS":
        return normalized

    if len(normalized) > NPDD_UPDATE_INDEX and str(normalized[NPDD_UPDATE_INDEX]).strip().upper() in {"#N/A", "N/A", "NA", "-"}:
        normalized[NPDD_UPDATE_INDEX] = ""
    if len(normalized) > NOW_KOL_INDEX:
        normalized[NOW_KOL_INDEX] = "Lunas"
    if len(normalized) > NOW_DETAIL_INDEX:
        normalized[NOW_DETAIL_INDEX] = "Lunas"

    for index in NOW_NUMERIC_INDEXES:
        if len(normalized) > index:
            normalized[index] = "0"

    if len(normalized) > PTP_INDEX:
        normalized[PTP_INDEX] = "LUNAS"

    return normalized


def detect_now_excel_offset(subheader_values):
    """Detect how many columns the NOW section is shifted right of index 21.

    NPDD Excel has a merged 'POSISI NOW' label cell at column 21; actual
    now_kol data lives one column further right (offset = 1).  We confirm
    this by locating the second 'KOL'-like sub-header after index 17.
    """
    normalized = [str(v or "").strip().upper().replace(" ", "_").replace(".", "_") for v in subheader_values]
    kol_like = {"KOL", "REF_KOL", "KOLEKTABILITAS"}
    kol_positions = [i for i, v in enumerate(normalized) if i >= 17 and v in kol_like]
    if len(kol_positions) >= 2:
        return max(0, kol_positions[1] - 21)
    if len(kol_positions) == 1:
        return max(0, kol_positions[0] - 21)
    return 0


def looks_like_npdd_header(row):
    values = [str(value or "").strip().upper() for value in row]

    return (
        "BILLING" in values
        and "PERIODE" in values
        and "KANCA" in values
        and "NOMOR REKENING" in values
        and "NPDD" in values
        and "NPDD UPDATE" in values
    )


def main():
    parser = argparse.ArgumentParser(description="Stream LW321 NPDD MICRO XLSX to normalized CSV staging.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output")
    parser.add_argument("--preview-only", action="store_true")
    parser.add_argument("--preview-limit", type=int, default=75)
    parser.add_argument("--progress-every", type=int, default=25000)
    args = parser.parse_args()

    if not args.preview_only and not args.output:
        raise RuntimeError("--output wajib diisi untuk mode staging penuh.")

    if args.output:
        os.makedirs(os.path.dirname(args.output), exist_ok=True)

    from openpyxl import load_workbook

    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    worksheet_max_row = ws.max_row or 0

    header_row = None
    now_excel_offset = 0
    rows_written = 0
    preview_rows = []
    unique_values = {}

    try:
        handle = None
        writer = None
        if not args.preview_only:
            handle = open(args.output, "w", newline="", encoding="utf-8")
            writer = csv.writer(handle)
            writer.writerow(HEADERS)

        try:
            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = list(row)

                if header_row is None:
                    if looks_like_npdd_header(values):
                        header_row = row_number
                        emit({
                            "type": "progress",
                            "percent": 45 if not args.preview_only else 65,
                            "message": "Header LW321 NPDD ditemukan. Menulis CSV staging..." if not args.preview_only else "Header LW321 NPDD ditemukan. Menyiapkan sampel preview...",
                            "header_row": header_row,
                            "headers": len(HEADERS),
                        })
                    continue

                if row_number == header_row + 1:
                    now_excel_offset = detect_now_excel_offset(values)
                    continue

                excel_values = values + [""] * (len(HEADERS) + now_excel_offset + 1)
                normalized = [
                    cell_to_csv_value(excel_values[i + now_excel_offset if i >= 21 else i], i)
                    for i in range(len(HEADERS))
                ]
                normalized = normalize_lunas_row(normalized)
                if not any(value for value in normalized):
                    continue

                if writer is not None:
                    writer.writerow(normalized)

                rows_written += 1

                if len(preview_rows) < args.preview_limit:
                    preview_rows.append(normalized)

                    for index, value in enumerate(normalized):
                        value = str(value).strip() or "(Blank)"
                        bucket = unique_values.setdefault(index, {})
                        if len(bucket) < 75:
                            bucket[value] = True

                if args.preview_only and len(preview_rows) >= args.preview_limit:
                    break

                if not args.preview_only and rows_written % max(1, args.progress_every) == 0:
                    emit({
                        "type": "progress",
                        "percent": min(78, 45 + int(rows_written / max(worksheet_max_row, 1) * 35)),
                        "message": f"Menulis CSV staging LW321 NPDD... {rows_written} baris",
                        "rows": rows_written,
                    })
        finally:
            if handle is not None:
                handle.close()
    finally:
        wb.close()

    if header_row is None:
        raise RuntimeError("Header LW321 NPDD tidak ditemukan. Pastikan file memuat Billing, KANCA, NOMOR REKENING, NPDD, dan NPDD UPDATE.")

    emit({
        "type": "done",
        "output": args.output if not args.preview_only else None,
        "header_index": 0,
        "source_header_row": header_row,
        "headers": HEADERS,
        "preview_rows": preview_rows,
        "unique_values": {str(index): list(values.keys()) for index, values in unique_values.items()},
        "total_rows": rows_written if not args.preview_only else max(0, worksheet_max_row - header_row - 1),
    })


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        emit({"type": "error", "message": str(exc)})
        sys.exit(1)
