import argparse
import csv
import json
import os
import re
import sys
import zipfile
from datetime import datetime, timedelta
from datetime import date
from xml.etree import ElementTree as ET

DATE_HEADERS = {
    "PERIODE",
    "NEXT_PMT_DATE",
    "NEXT_INT_PMT_DATE",
    "TGL_MENUNGGAK",
    "TGL_REALISASI",
    "TGL JATUH TEMPO",
}

SHARED_PREFIX = "__SST__"


def emit(payload):
    print(json.dumps(payload, ensure_ascii=False), flush=True)


def column_index(cell_ref):
    letters = re.sub(r"[^A-Z]", "", cell_ref.upper())
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - 64)
    return max(0, index - 1)


def excel_serial_to_date(value):
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return value

    if serial < 20000 or serial > 80000:
        return value

    # Excel's Windows date system includes the 1900 leap-year bug.
    date_value = datetime(1899, 12, 30) + timedelta(days=serial)
    return date_value.strftime("%d/%m/%Y")


def shared_string_at(shared_strings_path, wanted_indexes):
    if not wanted_indexes:
        return {}

    wanted = set(wanted_indexes)
    max_wanted = max(wanted)
    resolved = {}

    with zipfile.ZipFile(shared_strings_path.filename) as archive:
        with archive.open("xl/sharedStrings.xml") as handle:
            index = -1
            for event, elem in ET.iterparse(handle, events=("end",)):
                if elem.tag.endswith("}si"):
                    index += 1
                    if index in wanted:
                        texts = [text_node.text or "" for text_node in elem.iter() if text_node.tag.endswith("}t")]
                        resolved[index] = "".join(texts)
                    elem.clear()
                    if index >= max_wanted and wanted.issubset(resolved.keys()):
                        break

    return resolved


def read_cell_value(cell):
    cell_type = cell.attrib.get("t", "")

    if cell_type == "inlineStr":
        texts = [text_node.text or "" for text_node in cell.iter() if text_node.tag.endswith("}t")]
        return "".join(texts)

    value_node = None
    for child in cell:
        if child.tag.endswith("}v"):
            value_node = child
            break

    raw = "" if value_node is None or value_node.text is None else value_node.text
    if cell_type == "s" and raw != "":
        return f"{SHARED_PREFIX}{raw}"

    return raw


def resolve_shared_values(values, shared_map):
    resolved = []
    for value in values:
        if isinstance(value, str) and value.startswith(SHARED_PREFIX):
            try:
                resolved.append(shared_map.get(int(value[len(SHARED_PREFIX):]), ""))
            except ValueError:
                resolved.append("")
        else:
            resolved.append(value)
    return resolved


def fast_preview_xlsx(args):
    try:
        fast_preview_xlsx_fastexcel(args)
        return
    except ImportError:
        pass

    header_row = None
    headers = []
    preview_rows = []
    unique_values = {}
    shared_indexes = set()
    pending_rows = []

    with zipfile.ZipFile(args.input) as archive:
        sheet_name = "xl/worksheets/sheet1.xml"
        if sheet_name not in archive.namelist():
            sheet_name = next(
                (name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")),
                None,
            )
        if sheet_name is None:
            raise RuntimeError("Worksheet XLSX tidak ditemukan.")

        with archive.open(sheet_name) as handle:
            for event, row in ET.iterparse(handle, events=("end",)):
                if not row.tag.endswith("}row"):
                    continue

                row_number = int(row.attrib.get("r", "0") or 0)
                values = []
                for cell in row:
                    if not cell.tag.endswith("}c"):
                        continue
                    ref = cell.attrib.get("r", "")
                    index = column_index(ref)
                    while len(values) <= index:
                        values.append("")
                    value = read_cell_value(cell)
                    if isinstance(value, str) and value.startswith(SHARED_PREFIX):
                        try:
                            shared_indexes.add(int(value[len(SHARED_PREFIX):]))
                        except ValueError:
                            pass
                    values[index] = value

                pending_rows.append((row_number, values))
                row.clear()

                if header_row is not None and len(preview_rows) >= args.preview_limit:
                    break

                if header_row is None:
                    continue

                if len(preview_rows) < args.preview_limit and any(str(value).strip() for value in values):
                    preview_rows.append(values)

        shared_map = {}
        if "xl/sharedStrings.xml" in archive.namelist() and shared_indexes:
            with archive.open("xl/sharedStrings.xml") as handle:
                index = -1
                max_index = max(shared_indexes)
                for event, elem in ET.iterparse(handle, events=("end",)):
                    if elem.tag.endswith("}si"):
                        index += 1
                        if index in shared_indexes:
                            texts = [text_node.text or "" for text_node in elem.iter() if text_node.tag.endswith("}t")]
                            shared_map[index] = "".join(texts)
                        elem.clear()
                        if index >= max_index and shared_indexes.issubset(shared_map.keys()):
                            break

    resolved_rows = [(row_number, resolve_shared_values(values, shared_map)) for row_number, values in pending_rows]
    header_row = None
    preview_rows = []
    for row_number, values in resolved_rows:
        upper = [str(value).strip().upper() for value in values]
        if header_row is None:
            if "PERIODE" in upper and "NOMOR_REKENING" in upper:
                header_row = row_number
                headers = [
                    str(value).strip() if value and str(value).strip() else f"COL_{index}"
                    for index, value in enumerate(values)
                ]
                emit({
                    "type": "progress",
                    "percent": 65,
                    "message": "Header LW321PN ditemukan. Menyiapkan sampel preview...",
                    "header_row": header_row,
                    "headers": len(headers),
                })
            continue

        if len(preview_rows) >= args.preview_limit:
            break
        if any(str(value).strip() for value in values):
            normalized = (values + [""] * len(headers))[:len(headers)]
            for index, header in enumerate(headers):
                if str(header).strip().upper() in DATE_HEADERS:
                    normalized[index] = excel_serial_to_date(normalized[index])
            preview_rows.append(normalized)

    if header_row is None:
        raise RuntimeError("Header LW321PN tidak ditemukan. Pastikan file memuat PERIODE dan NOMOR_REKENING.")

    for normalized in preview_rows:
        for index, value in enumerate(normalized):
            value = str(value).strip()
            if value == "":
                value = "(Blank)"
            bucket = unique_values.setdefault(index, {})
            if len(bucket) < 75:
                bucket[value] = True

    emit({
        "type": "done",
        "output": None,
        "header_index": 0,
        "source_header_row": header_row,
        "headers": headers,
        "preview_rows": preview_rows,
        "unique_values": {str(index): list(values.keys()) for index, values in unique_values.items()},
        "total_rows": len(preview_rows),
    })


def preview_string(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def fast_preview_xlsx_fastexcel(args):
    import fastexcel

    scan_rows = max(200, args.preview_limit + 20)
    reader = fastexcel.read_excel(args.input)
    sheet = reader.load_sheet_by_idx(
        0,
        header_row=None,
        n_rows=scan_rows,
        schema_sample_rows=min(scan_rows, 100),
        dtype_coercion="coerce",
    )
    frame = sheet.to_polars()
    rows = [list(row) for row in frame.iter_rows()]

    header_offset = None
    headers = []
    for index, row in enumerate(rows):
        values = [preview_string(value) for value in row]
        upper = [value.strip().upper() for value in values]
        if "PERIODE" in upper and "NOMOR_REKENING" in upper:
            header_offset = index
            headers = [
                value.strip() if value and value.strip() else f"COL_{column_index}"
                for column_index, value in enumerate(values)
            ]
            break

    if header_offset is None:
        raise RuntimeError("Header LW321PN tidak ditemukan. Pastikan file memuat PERIODE dan NOMOR_REKENING.")

    emit({
        "type": "progress",
        "percent": 65,
        "message": "Header LW321PN ditemukan. Menyiapkan sampel preview...",
        "header_row": header_offset + 1,
        "headers": len(headers),
    })

    preview_rows = []
    unique_values = {}
    for row in rows[header_offset + 1:]:
        if len(preview_rows) >= args.preview_limit:
            break

        normalized = [preview_string(value) for value in row]
        normalized = (normalized + [""] * len(headers))[:len(headers)]
        if not any(str(value).strip() for value in normalized):
            continue

        for index, header in enumerate(headers):
            if str(header).strip().upper() in DATE_HEADERS:
                normalized[index] = excel_serial_to_date(normalized[index])

        preview_rows.append(normalized)

        for index, value in enumerate(normalized):
            value = str(value).strip()
            if value == "":
                value = "(Blank)"
            bucket = unique_values.setdefault(index, {})
            if len(bucket) < 75:
                bucket[value] = True

    emit({
        "type": "done",
        "output": None,
        "header_index": 0,
        "source_header_row": header_offset + 1,
        "headers": headers,
        "preview_rows": preview_rows,
        "unique_values": {str(index): list(values.keys()) for index, values in unique_values.items()},
        "total_rows": max(0, int(sheet.total_height or 0) - (header_offset + 1)),
    })


def main():
    parser = argparse.ArgumentParser(description="Stream LW321PN XLSX to CSV staging.")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output")
    parser.add_argument("--preview-only", action="store_true")
    parser.add_argument("--preview-limit", type=int, default=75)
    parser.add_argument("--progress-every", type=int, default=25000)
    args = parser.parse_args()

    if not args.preview_only:
        if not args.output:
            raise RuntimeError("--output wajib diisi untuk mode staging penuh.")
        os.makedirs(os.path.dirname(args.output), exist_ok=True)
    else:
        fast_preview_xlsx(args)
        return

    from openpyxl import load_workbook
    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    worksheet_max_row = ws.max_row or 0

    header_row = None
    headers = []
    rows_written = 0
    preview_rows = []
    unique_values = {}

    try:
        handle = None
        writer = None
        if not args.preview_only:
            handle = open(args.output, "w", newline="", encoding="utf-8")
            writer = csv.writer(handle)

        try:
            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                values = ["" if value is None else str(value) for value in row]
                upper = [value.strip().upper() for value in values]

                if header_row is None:
                    if "PERIODE" in upper and "NOMOR_REKENING" in upper:
                        header_row = row_number
                        headers = [
                            value.strip() if value and value.strip() else f"COL_{index}"
                            for index, value in enumerate(values)
                        ]
                        if writer is not None:
                            writer.writerow(headers)
                        emit({
                            "type": "progress",
                            "percent": 45 if not args.preview_only else 65,
                            "message": "Header LW321PN ditemukan. Menulis CSV staging..." if not args.preview_only else "Header LW321PN ditemukan. Menyiapkan sampel preview...",
                            "header_row": header_row,
                            "headers": len(headers),
                        })
                    continue

                if not any(str(value).strip() for value in values):
                    continue

                normalized = (values + [""] * len(headers))[:len(headers)]
                if writer is not None:
                    writer.writerow(normalized)
                rows_written += 1

                if len(preview_rows) < args.preview_limit:
                    preview_rows.append(normalized)

                if rows_written <= args.preview_limit:
                    for index, value in enumerate(normalized):
                        value = str(value).strip()
                        if value == "":
                            value = "(Blank)"
                        bucket = unique_values.setdefault(index, {})
                        if len(bucket) < 75:
                            bucket[value] = True

                if args.preview_only and len(preview_rows) >= args.preview_limit:
                    break

                if not args.preview_only and rows_written % max(1, args.progress_every) == 0:
                    emit({
                        "type": "progress",
                        "percent": min(78, 45 + int(rows_written / max(ws.max_row or rows_written, 1) * 35)),
                        "message": f"Menulis CSV staging LW321PN... {rows_written} baris",
                        "rows": rows_written,
                    })
        finally:
            if handle is not None:
                handle.close()
    finally:
        wb.close()

    if header_row is None:
        raise RuntimeError("Header LW321PN tidak ditemukan. Pastikan file memuat PERIODE dan NOMOR_REKENING.")

    emit({
        "type": "done",
        "output": args.output if not args.preview_only else None,
        "header_index": 0,
        "source_header_row": header_row,
        "headers": headers,
        "preview_rows": preview_rows,
        "unique_values": {str(index): list(values.keys()) for index, values in unique_values.items()},
        "total_rows": max(0, worksheet_max_row - header_row) if args.preview_only else rows_written,
    })


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        emit({"type": "error", "message": str(exc)})
        sys.exit(1)
