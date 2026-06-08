#!/usr/bin/env python3
"""
Excel CPU Processor untuk Laravel Import
=========================================
Arsitektur:
  - Python  : baca Excel dengan pandas (CPU), normalisasi, output batch JSON ke stdout
  - PHP     : terima batch JSON dari stdout, insert ke database (tidak perlu pymysql)

Dependencies:
  pip install polars pandas openpyxl python-dateutil
"""

import sys
import json
import os
import argparse
import time
import uuid
import csv
import re
from datetime import datetime, date, timedelta

# ── Force CPU: matikan semua GPU device sebelum import apapun ────────────────
os.environ['CUDA_VISIBLE_DEVICES']   = ''
os.environ['ROCR_VISIBLE_DEVICES']   = ''
os.environ['MLU_VISIBLE_DEVICES']    = ''
os.environ['ASCEND_VISIBLE_DEVICES'] = ''
os.environ['HIP_VISIBLE_DEVICES']    = ''


# ─────────────────────────────────────────────────────────────────────────────
# Helper: kirim event ke PHP (stdout)
# ─────────────────────────────────────────────────────────────────────────────

def send_event(event_type, data):
    data['type'] = event_type
    print(json.dumps(data, ensure_ascii=False, default=str), flush=True)


def send_progress(percent, message, rows_done=0, total=0, speed=0):
    send_event('progress', {
        'percent':   percent,
        'message':   message,
        'rows_done': rows_done,
        'total':     total,
        'speed':     speed,
    })


def send_error(message):
    send_event('error', {'message': message})


def _is_empty_row(values):
    for value in values:
        if value is None:
            continue
        if str(value).strip() != '':
            return False
    return True


def _read_excel_with_openpyxl(file_path, header_index):
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row_excel = header_index + 1
        headers = []
        row_values = []

        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_list = list(row)

            if idx == header_row_excel:
                headers = [str(v).strip() if v is not None and str(v).strip() != '' else 'COL_' + str(i) for i, v in enumerate(row_list)]
                continue

            if idx <= header_row_excel:
                continue

            if _is_empty_row(row_list):
                continue

            row_values.append(row_list)

        return headers, row_values
    finally:
        wb.close()


def read_excel_table(file_path, header_index, preserve_column_positions=False):
    """
    Try Polars first for lower memory + multithreaded parse.
    Fallback to pandas to keep compatibility on older environments.
    Returns: (headers, row_values, backend_name)
    """
    if preserve_column_positions:
        headers, row_values = _read_excel_with_openpyxl(file_path, header_index)
        return headers, row_values, 'openpyxl-column-preserving'

    polars_errors = []

    try:
        import polars as pl
        read_excel = getattr(pl, 'read_excel', None)
        if callable(read_excel):
            # Kombinasi opsi untuk kompatibilitas lintas versi Polars.
            read_attempts = [
                lambda: read_excel(source=file_path, sheet_id=1, engine='calamine', read_options={'header_row': header_index}),
                lambda: read_excel(source=file_path, sheet_id=1, read_options={'header_row': header_index}),
                lambda: read_excel(file_path, sheet_id=1, engine='calamine', read_options={'header_row': header_index}),
                lambda: read_excel(file_path, sheet_id=1, read_options={'header_row': header_index}),
                lambda: read_excel(source=file_path, sheet_id=1, engine='calamine', read_options={'has_header': True, 'skip_rows': header_index}),
                lambda: read_excel(file_path, sheet_id=1, engine='calamine', read_options={'has_header': True, 'skip_rows': header_index}),
            ]

            for attempt in read_attempts:
                try:
                    df_pl = attempt()
                    headers = [str(col) for col in df_pl.columns]
                    row_values = [list(row) for row in df_pl.rows() if not _is_empty_row(row)]
                    return headers, row_values, 'polars'
                except Exception as e:
                    polars_errors.append(type(e).__name__ + ': ' + str(e))
                    continue
    except Exception as e:
        polars_errors.append(type(e).__name__ + ': ' + str(e))

    try:
        import pandas as pd
        df_pd = pd.read_excel(
            file_path,
            header=header_index,
            engine='openpyxl',
            dtype=object,
        )
        df_pd = df_pd.dropna(how='all').reset_index(drop=True)
        headers = [str(col) for col in list(df_pd.columns)]
        row_values = df_pd.values.tolist()

        if polars_errors:
            return headers, row_values, 'pandas-fallback'

        return headers, row_values, 'pandas'
    except Exception:
        pass

    headers, row_values = _read_excel_with_openpyxl(file_path, header_index)
    if polars_errors:
        return headers, row_values, 'openpyxl-fallback'
    return headers, row_values, 'openpyxl'


# ─────────────────────────────────────────────────────────────────────────────
# Normalisasi nilai sel Excel
# ─────────────────────────────────────────────────────────────────────────────

EXCEL_EPOCH  = date(1899, 12, 30)
DATE_COLUMNS = {'PERIODE', 'POSISI', 'MONTH_DAY_YEAR_OF_POSISI', 'MONTH_DAY_YEAR_OF_PERIODE', 'TGL_REALISASI', 'TGL_JATUH_TEMPO', 'TANGGAL'}
DECIMAL_COLUMNS = {
    'BAKI_DEBET', 'SALDO', 'POKOK', 'BUNGA', 'PLAFON', 'BESAR_REALISASI',
    'ANGPOK', 'ANGBUNG', 'SISAPOK', 'SISABUN', 'OS_PENUH_BERJALAN',
    'SALDO_PERTAMA_PH_POKOK', 'SALDO_PERTAMA_PH_BUNGA', 'ANGR_POKOK',
    'ANGR_BUNGA', 'OS', 'TOTAL_FEE', 'TOTAL_NOMINAL', 'JUMLAH', 'NILAI',
    'BAKI_DEBET1', 'CKPN', 'BAP', 'BILPRN', 'BILINT', 'BILLC', 'PMTAMT',
    'TUNGGAKAN_POKOK', 'TUNGGAKAN_BUNGA'
}
SIMPANAN_DECIMAL_COLUMNS = {'SALDO_IDR'}
NULL_STRS    = {'', 'nan', 'none', 'nat', 'null', 'n/a', 'na'}
INDONESIAN_MONTHS = {
    'januari': 'january',
    'februari': 'february',
    'maret': 'march',
    'april': 'april',
    'mei': 'may',
    'juni': 'june',
    'juli': 'july',
    'agustus': 'august',
    'september': 'september',
    'oktober': 'october',
    'november': 'november',
    'desember': 'december',
}
LOCALE_DATE_TABLES = {'hourly_dpk', 'ssa_simpanan', 'ssa_pinjaman'}


def canonicalize_header(header_name):
    import re
    return re.sub(r'[^A-Z0-9]+', '_', str(header_name).upper().strip()).strip('_')


def normalize_locale_date_text(value: str) -> str:
    normalized = value.strip()
    for source, target in INDONESIAN_MONTHS.items():
        normalized = re.sub(rf'\b{source}\b', target, normalized, flags=re.IGNORECASE)
    return normalized


def allows_locale_date_text(table_name) -> bool:
    return str(table_name or '').strip().lower() in LOCALE_DATE_TABLES


def is_simpanan_multipn_table(table_name) -> bool:
    return str(table_name or '').strip().lower() == 'simpanan_multipn'


def is_gi405_recovery_table(table_name) -> bool:
    return str(table_name or '').strip().lower() == 'gi405_recovery'


def normalize_decimal_value(value):
    if value is None:
        return None

    value_str = str(value).strip()
    if value_str.lower() in NULL_STRS:
        return None

    value_str = ''.join(value_str.split())
    filtered = ''.join(ch for ch in value_str if ch.isdigit() or ch in ',.-')

    if filtered in ('', '-', None):
        return None

    has_comma = ',' in filtered
    has_dot = '.' in filtered

    if has_comma and has_dot:
        if filtered.rfind(',') > filtered.rfind('.'):
            filtered = filtered.replace('.', '')
            filtered = filtered.replace(',', '.')
        else:
            filtered = filtered.replace(',', '')
    elif has_comma:
        parts = filtered.split(',')
        last_part = parts[-1]
        if len(parts) > 2 or len(last_part) == 3:
            filtered = filtered.replace(',', '')
        else:
            filtered = filtered.replace(',', '.')
    elif has_dot:
        parts = filtered.split('.')
        last_part = parts[-1]
        if len(parts) > 2 or len(last_part) == 3:
            filtered = filtered.replace('.', '')

    try:
        return '{:.2f}'.format(float(filtered))
    except (ValueError, TypeError):
        return None


def normalize_value(header_name, value, table_name=None):
    import math
    header = canonicalize_header(header_name)

    if value is None:
        return None

    if isinstance(value, float) and math.isnan(value):
        return None

    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d') if header in DATE_COLUMNS else value.strftime('%Y-%m-%d %H:%M:%S')

    value_str = str(value).strip()
    if value_str.lower() in NULL_STRS:
        return None

    if header in DATE_COLUMNS:
        try:
            try:
                num = float(value_str)
                d = EXCEL_EPOCH + timedelta(days=int(num))
                return d.strftime('%Y-%m-%d')
            except (ValueError, OverflowError):
                pass
            if is_gi405_recovery_table(table_name):
                match = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?$', value_str)
                if match:
                    day, month, year = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
                    return date(year, month, day).strftime('%Y-%m-%d')
            from dateutil import parser as dateutil_parser
            normalized_date_text = value_str
            if allows_locale_date_text(table_name):
                normalized_date_text = normalize_locale_date_text(normalized_date_text)
            normalized_date_text = normalized_date_text.replace('/', '-')
            return dateutil_parser.parse(normalized_date_text).strftime('%Y-%m-%d')
        except Exception:
            return None

    if header in DECIMAL_COLUMNS or (is_simpanan_multipn_table(table_name) and header in SIMPANAN_DECIMAL_COLUMNS):
        return normalize_decimal_value(value)

    # Fallback: jika nilainya terlihat seperti angka dengan ribuan (misal "219,000.00")
    if isinstance(value_str, str) and ',' in value_str and any(c.isdigit() for c in value_str):
        normalized = normalize_decimal_value(value_str)
        if normalized is not None:
            return normalized

    try:
        num = float(value_str)
        formatted = '{:.2f}'.format(num).rstrip('0').rstrip('.')
        return formatted if formatted != '' else '0'
    except (ValueError, TypeError):
        pass

    return value_str


def is_valid_simpanan_import_row(final_row):
    posisi = final_row.get('posisi')
    cifno = str(final_row.get('CIFNO') if 'CIFNO' in final_row else final_row.get('cifno') or '').strip()
    no_rekening = str(final_row.get('no_rekening') or '').strip()
    jenis = str(final_row.get('jenis_simpanan') or '').strip().upper()
    saldo = final_row.get('saldo_idr')

    if posisi is None or str(posisi).strip() == '':
        return False
    if cifno == '' or no_rekening == '' or jenis == '' or saldo is None or str(saldo).strip() == '':
        return False
    if re.fullmatch(r"[A-Z0-9.,+_\/'-]+", no_rekening, flags=re.I) is None:
        return False
    if len(no_rekening) < 6:
        return False
    if not (jenis.startswith('TABUNGAN') or jenis.startswith('GIRO') or jenis.startswith('DEPOSITO')):
        return False

    return normalize_decimal_value(saldo) is not None


# ─────────────────────────────────────────────────────────────────────────────
# MODE: init — Scan cepat header & total baris
# ─────────────────────────────────────────────────────────────────────────────

def run_init(config):
    file_path = config['file_path']

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        scan_rows = []
        for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
            scan_rows.append(list(row))
    except Exception as e:
        print(json.dumps({'status': 'error', 'message': 'Gagal membuka file: ' + str(e)}), flush=True)
        sys.exit(1)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    header_index  = None
    header_values = []

    best_row_index = None
    best_row_values = []
    best_score = -1

    for i in range(len(scan_rows)):
        row       = scan_rows[i]
        row_upper = [str(v).upper().strip() if str(v).lower() not in ('nan', 'none', '') else '' for v in row]
        filled_cells = sum(1 for value in row_upper if value != '')
        contains_posisi = any('POSISI' in value for value in row_upper)
        contains_periode = any('PERIODE' in value for value in row_upper)

        if filled_cells > best_score:
            best_score = filled_cells
            best_row_index = i
            best_row_values = [str(v).strip() if str(v).lower() not in ('nan', 'none') else '' for v in row]

        if contains_periode or contains_posisi:
            header_index  = i
            header_values = [str(v).strip() if str(v).lower() not in ('nan', 'none') else '' for v in row]
            break

    if header_index is None:
        if best_row_index is not None and best_score >= 3:
            header_index = best_row_index
            header_values = best_row_values
        else:
            print(json.dumps({
                'status':  'error',
                'message': 'Header utama tidak ditemukan dalam 200 baris pertama.',
            }), flush=True)
            sys.exit(1)

    # Total baris via openpyxl read-only (cepat, dari metadata XML)
    total_rows = 0
    try:
        from openpyxl import load_workbook
        wb         = load_workbook(file_path, read_only=True, data_only=True)
        total_rows = wb.active.max_row or 0
        wb.close()
    except Exception:
        total_rows = 0

    print(json.dumps({
        'status':        'ok',
        'header_index':  header_index,
        'total_rows':    total_rows,
        'header_values': header_values,
    }), flush=True)
    sys.exit(0)


# ─────────────────────────────────────────────────────────────────────────────
# MODE: process — Baca Excel dengan pandas, output batch JSON ke stdout
#                 PHP yang akan insert ke database (tidak perlu pymysql)
# ─────────────────────────────────────────────────────────────────────────────

def run_process(config):
    try:
        _run_process_inner(config)
    except Exception as e:
        # Tangkap semua exception yang tidak tertangani dan kirim sebagai error event
        # agar PHP bisa fallback ke chunked reading
        import traceback
        send_error('Python error: ' + str(e) + ' | ' + traceback.format_exc(limit=3))
        sys.exit(1)


def _run_process_inner(config):
    file_path          = config['file_path']
    header_index       = int(config['header_index'])
    table_name         = config['table_name']
    active_filters     = config.get('active_filters', {})
    normalized_headers = config['normalized_headers']
    table_columns_raw  = [str(c).strip() for c in config.get('table_columns', []) if str(c).strip() != '']
    table_columns      = set(c.lower() for c in table_columns_raw)

    # ── CRITICAL FIX: PHP json_encode mengubah array integer-key menjadi JSON array ──
    # Contoh: [0=>'PERIODE', 1=>'POSISI'] → ["PERIODE","POSISI"] (bukan {"0":"PERIODE",...})
    # Python menerima list, bukan dict → normalized_headers.keys() crash!
    # Fix: konversi list ke dict dengan index sebagai key string
    if isinstance(normalized_headers, list):
        normalized_headers = {str(i): v for i, v in enumerate(normalized_headers)}

    # ── Baca seluruh file dengan pandas (CPU, satu kali load) ────────────────
    send_progress(5, 'Membaca file Excel (Polars preferred, pandas fallback)...')

    try:
        preserve_column_positions = bool(config.get('preserve_column_positions')) or is_simpanan_multipn_table(table_name) or is_gi405_recovery_table(table_name)
        _headers, row_values, backend = read_excel_table(file_path, header_index, preserve_column_positions)
        total_rows = len(row_values)
        send_progress(20, 'File dibaca via ' + backend + ': ' + str(total_rows) + ' baris. Memproses kolom...')
    except Exception as e:
        send_error('Gagal membaca file Excel: ' + str(e))
        sys.exit(1)

    is_simpanan_multipn = is_simpanan_multipn_table(table_name)
    unique_id_col = str(config.get('unique_id_col') or ('uniqueid_SimoPN' if is_simpanan_multipn else 'uniqueid_namareport')).strip()
    suffix        = str(config.get('unique_id_suffix') if config.get('unique_id_suffix') is not None else ('_SimoPN' if is_simpanan_multipn else '_DLD'))
    table_columns_map = {str(col).lower(): str(col) for col in table_columns_raw}
    unique_id_col = table_columns_map.get(unique_id_col.lower(), unique_id_col)
    unique_id_prefix = str(config.get('unique_id_prefix') or 'imp').strip() or 'imp'
    skip_cols     = set(['id', unique_id_col.lower()])
    manual_values = {}
    for manual_col, manual_value in dict(config.get('manual_values') or {}).items():
        manual_col_key = str(manual_col).strip()
        if manual_col_key == '':
            continue
        manual_col_lower = manual_col_key.lower()
        if table_columns and manual_col_lower not in table_columns_map:
            continue
        manual_values[table_columns_map.get(manual_col_lower, manual_col_key)] = manual_value

    # Build valid headers list: [(original_col_index, header_name), ...]
    valid_headers = []
    for idx_str in sorted(normalized_headers.keys(), key=lambda x: int(x)):
        h = normalized_headers[idx_str]
        if not h.startswith('COL_'):
            valid_headers.append((int(idx_str), h))

    output_csv_path = config.get('output_csv_path')
    load_columns = [str(col).strip() for col in config.get('load_columns', []) if str(col).strip() != '']

    if output_csv_path and not load_columns:
        send_error('Kolom bulk load tidak tersedia untuk export CSV.')
        sys.exit(1)

    send_progress(25, 'Mapping kolom selesai. Menyiapkan output import...', 0, total_rows)

    now_str    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    batch      = []
    batch_size = 200    # 200 baris per batch JSON line (aman untuk max_allowed_packet)
    rows_done  = 0
    start_time = time.time()

    csv_file = None
    csv_writer = None
    if output_csv_path:
        os.makedirs(os.path.dirname(output_csv_path), exist_ok=True)
        csv_file = open(output_csv_path, 'w', encoding='utf-8', newline='')
        csv_writer = csv.writer(
            csv_file,
            delimiter=',',
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL,
            lineterminator='\n',
            doublequote=True,
        )

    for row_list in row_values:
        mapped_data = {}
        pass_filter = True

        for filter_idx, (original_index, h_name) in enumerate(valid_headers):
            val = row_list[original_index] if original_index < len(row_list) else None
            val = normalize_value(h_name, val, table_name)

            filter_key = str(filter_idx)
            if active_filters and filter_key in active_filters:
                f_val = '(Blank)' if val is None else str(val)
                if f_val not in active_filters[filter_key]:
                    pass_filter = False
                    break

            mapped_data[h_name.upper().replace(' ', '_')] = val

        if not pass_filter:
            continue

        # Bangun baris final dengan unique ID dan timestamps
        final_row = {
            unique_id_col: unique_id_prefix + '_' + str(uuid.uuid4()) + suffix,
            'created_at':  now_str,
            'updated_at':  now_str,
        }
        for excel_key, val in mapped_data.items():
            db_col = excel_key.lower()
            if db_col in skip_cols:
                continue
            if table_columns and db_col not in table_columns_map:
                continue
            final_row[table_columns_map.get(db_col, db_col)] = val

        for manual_col, manual_value in manual_values.items():
            if manual_col.lower() in skip_cols:
                continue
            final_row[manual_col] = manual_value

        if is_simpanan_multipn and not is_valid_simpanan_import_row(final_row):
            continue

        if len(final_row) > 3:
            rows_done += 1

            if csv_writer is not None:
                csv_writer.writerow([
                    r'\N' if final_row.get(column) is None else final_row.get(column)
                    for column in load_columns
                ])
            else:
                batch.append(final_row)

        if csv_writer is None and len(batch) >= batch_size:
            print(json.dumps({'type': 'batch', 'rows': batch}, ensure_ascii=False, default=str), flush=True)
            batch = []

        # Kirim progress setiap 5000 baris
        if rows_done > 0 and rows_done % 5000 == 0:
            elapsed = max(time.time() - start_time, 0.001)
            speed   = int(rows_done / elapsed)
            pct     = min(90, 25 + int((rows_done / total_rows) * 65)) if total_rows > 0 else 50
            send_progress(pct, 'Memproses... (' + str(speed) + ' baris/detik)', rows_done, total_rows, speed)

    if csv_file is not None:
        if is_simpanan_multipn and rows_done == 0:
            csv_file.close()
            send_error('Tidak ada baris Simpanan MultiPN valid. Periksa mapping kolom Excel sebelum import.')
            sys.exit(1)
        csv_file.close()
        send_progress(95, 'CSV sementara selesai dibuat. Menunggu MySQL bulk load...', rows_done, total_rows)
        send_event('done', {'total_rows': rows_done, 'csv_path': output_csv_path})
        return

    if batch:
        print(json.dumps({'type': 'batch', 'rows': batch}, ensure_ascii=False, default=str), flush=True)

    if is_simpanan_multipn and rows_done == 0:
        send_error('Tidak ada baris Simpanan MultiPN valid. Periksa mapping kolom Excel sebelum import.')
        sys.exit(1)

    send_progress(95, 'File selesai diproses. Menunggu PHP selesai insert ke database...', rows_done, total_rows)
    send_event('done', {'total_rows': rows_done})


def run_stage(config):
    file_path = config['file_path']
    header_index = int(config['header_index'])
    normalized_headers = config['normalized_headers']
    output_csv_path = config['output_csv_path']
    table_name = config.get('table_name')

    if isinstance(normalized_headers, list):
        normalized_headers = {str(i): v for i, v in enumerate(normalized_headers)}

    try:
        preserve_column_positions = bool(config.get('preserve_column_positions')) or is_simpanan_multipn_table(table_name) or is_gi405_recovery_table(table_name)
        _headers, row_values, backend = read_excel_table(file_path, header_index, preserve_column_positions)
    except Exception as e:
        send_error('Gagal membaca file Excel untuk staging: ' + str(e))
        sys.exit(1)

    valid_headers = []
    for idx_str in sorted(normalized_headers.keys(), key=lambda x: int(x)):
        h = normalized_headers[idx_str]
        if not str(h).startswith('COL_'):
            valid_headers.append((int(idx_str), str(h)))

    total_rows = len(row_values)
    send_progress(10, 'Memulai konversi Excel ke CSV stage via ' + backend + '...', 0, total_rows, 0)

    os.makedirs(os.path.dirname(output_csv_path), exist_ok=True)

    rows_done = 0
    start_time = time.time()

    with open(output_csv_path, 'w', encoding='utf-8', newline='') as csv_file:
        writer = csv.writer(
            csv_file,
            delimiter=',',
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL,
            lineterminator='\n',
            doublequote=True,
        )

        writer.writerow([header_name for _, header_name in valid_headers])

        for row_list in row_values:
            output_row = []
            has_value = False

            for original_index, header_name in valid_headers:
                val = row_list[original_index] if original_index < len(row_list) else None
                val = normalize_value(header_name, val, table_name)
                if val is not None and str(val).strip() != '':
                    has_value = True
                output_row.append(r'\N' if val is None else val)

            if not has_value:
                continue

            writer.writerow(output_row)
            rows_done += 1

            if rows_done > 0 and rows_done % 5000 == 0:
                elapsed = max(time.time() - start_time, 0.001)
                speed = int(rows_done / elapsed)
                pct = min(95, 10 + int((rows_done / max(total_rows, 1)) * 80))
                send_progress(pct, 'Mengonversi ke CSV stage... (' + str(speed) + ' baris/detik)', rows_done, total_rows, speed)

    send_progress(98, 'CSV stage selesai dibuat.', rows_done, total_rows, 0)
    send_event('done', {'total_rows': rows_done, 'csv_path': output_csv_path})


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Excel CPU Processor untuk Laravel Import')
    parser.add_argument('--config', required=True, help='Path ke file config JSON')
    parser.add_argument('--mode',   default='process', choices=['init', 'process', 'stage'],
                        help='Mode: init | process | stage')
    args = parser.parse_args()

    try:
        with open(args.config, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except Exception as e:
        if args.mode == 'init':
            print(json.dumps({'status': 'error', 'message': 'Gagal membaca config: ' + str(e)}), flush=True)
        else:
            send_error('Gagal membaca config: ' + str(e))
        sys.exit(1)

    if args.mode == 'init':
        run_init(config)
    elif args.mode == 'stage':
        run_stage(config)
    else:
        run_process(config)


if __name__ == '__main__':
    main()
